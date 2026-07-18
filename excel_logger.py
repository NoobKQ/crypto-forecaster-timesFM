import os
import re
import logging
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

# Column layout:
# A: ID, B: ForecastTime, C: Coin, D: Timeframe, E: Horizon,
# F: LastPredicted, G: LastActual, H: AbsError, I: PctError, J: MAPE,
# K: Chart (image)

HEADERS = [
    "ID", "ForecastTime", "Coin", "Timeframe", "Horizon",
    "LastPredicted", "LastActual", "AbsError", "PctError", "MAPE"
]
IMAGE_COL = "K"

DETAIL_HEADERS = [
    "ID", "ForecastTime", "Coin", "Timeframe", "Horizon",
    "ForecastedPrices", "ActualPrices", "AbsError", "PctError", "MAPE",
    "LowerBand", "MedianBand", "UpperBand"
]

class ExcelLogger:
    def __init__(self, model_name="timesfm-2.5-200m-pytorch", coin="BTC-USD", timeframe="1h", horizon=24, period="30d"):
        sanitized = re.sub(r'[\\/:*?"<>|]', '_', os.path.basename(model_name))
        self.base_dir = os.path.join(sanitized, f"{coin}_{timeframe}_{period}_h{horizon}")
        self.charts_dir = os.path.join(self.base_dir, "charts")
        self.log_filepath = os.path.join(self.base_dir, "log.xlsx")

        os.makedirs(self.base_dir, exist_ok=True)
        os.makedirs(self.charts_dir, exist_ok=True)

    def _get_or_create_workbook(self):
        if not os.path.exists(self.log_filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "Forecast Logs"
            ws.append(HEADERS)
            ws2 = wb.create_sheet("Forecast Details")
            ws2.append(DETAIL_HEADERS)
            wb.save(self.log_filepath)
        return load_workbook(self.log_filepath)

    def _get_next_id_and_row(self, ws):
        max_row = ws.max_row
        while max_row > 1:
            row_vals = [ws.cell(row=max_row, column=c).value for c in range(1, len(HEADERS) + 1)]
            if any(v is not None for v in row_vals):
                break
            max_row -= 1

        next_row = max_row + 1

        if max_row == 1:
            next_id = 1
        else:
            last_id = ws.cell(row=max_row, column=1).value
            try:
                next_id = int(last_id) + 1
            except (ValueError, TypeError):
                next_id = 1

        return next_id, next_row

    def log_forecast(self, coin, timeframe, horizon, forecasted,
                     lower_band=None, median_band=None, upper_band=None):
        """
        Logs a new future forecast row. Actuals/errors/chart are left blank.
        """
        forecasted = np.array(forecasted).flatten()

        wb = self._get_or_create_workbook()
        ws = wb["Forecast Logs"]
        ws_detail = wb["Forecast Details"]
        next_id, next_row = self._get_next_id_and_row(ws)

        now_time = pd.Timestamp.now(tz='UTC')
        forecast_time_str = now_time.strftime("%Y-%m-%d %H:%M:%S")

        last_predicted = float(forecasted[-1])
        forecasted_seq_str = ",".join(f"{v:.6f}" for v in forecasted)

        # Write to Forecast Logs sheet
        data_row = [
            next_id,
            forecast_time_str,
            coin,
            timeframe,
            horizon,
            last_predicted,
            None,  # LastActual
            None,  # AbsError
            None,  # PctError
            None,  # MAPE
        ]
        for col_idx, val in enumerate(data_row, start=1):
            ws.cell(row=next_row, column=col_idx, value=val)

        # Write to ForecastDetails sheet
        detail_headers = [h.value for h in ws_detail[1]]
        detail_row = [None] * len(detail_headers)
        detail_row[detail_headers.index("ID")]             = next_id
        detail_row[detail_headers.index("ForecastTime")]   = forecast_time_str
        detail_row[detail_headers.index("Coin")]           = coin
        detail_row[detail_headers.index("Timeframe")]      = timeframe
        detail_row[detail_headers.index("Horizon")]        = horizon
        detail_row[detail_headers.index("ForecastedPrices")] = forecasted_seq_str

        lower_arr = np.array(lower_band).flatten() if lower_band is not None else None
        median_arr = np.array(median_band).flatten() if median_band is not None else None
        upper_arr = np.array(upper_band).flatten() if upper_band is not None else None

        detail_row[detail_headers.index("LowerBand")]  = ",".join(f"{v:.6f}" for v in lower_arr) if lower_arr is not None else None
        detail_row[detail_headers.index("MedianBand")] = ",".join(f"{v:.6f}" for v in median_arr) if median_arr is not None else None
        detail_row[detail_headers.index("UpperBand")]  = ",".join(f"{v:.6f}" for v in upper_arr) if upper_arr is not None else None

        next_detail_row = ws_detail.max_row + 1
        for col_idx, val in enumerate(detail_row, start=1):
            ws_detail.cell(row=next_detail_row, column=col_idx, value=val)

        wb.save(self.log_filepath)
        logging.info(f"Logged new future forecast ID {next_id} at row {next_row}. Pending actuals.")
        return next_id

    def check_and_evaluate_latest(self, data):
        """
        Check the latest forecast row. If it has no actual_price, read the
        ForecastTime from Excel, round it down to the interval boundary, add
        horizon * interval, then find that exact timestamp in the downloaded
        data to get the actual price. Calculates metrics and writes them.
        Returns dict with evaluated=True + metrics on success, empty dict if skipped.
        """
        if not os.path.exists(self.log_filepath):
            logging.info("No log file found. Skipping evaluation.")
            return {}

        wb = load_workbook(self.log_filepath)
        ws = wb.active

        latest_row = None
        for r in range(ws.max_row, 1, -1):
            if ws.cell(row=r, column=1).value is not None:
                latest_row = r
                break

        if latest_row is None:
            logging.info("No forecast rows found. Skipping evaluation.")
            return {}

        if ws.cell(row=latest_row, column=7).value is not None:
            logging.info(f"Latest forecast (row {latest_row}) already has actual_price. Skipping evaluation.")
            return {}

        logging.info(f"Latest forecast (row {latest_row}) is pending. Evaluating...")
        r = latest_row

        row_id             = ws.cell(row=r, column=1).value
        forecast_time_val  = ws.cell(row=r, column=2).value
        coin               = ws.cell(row=r, column=3).value
        tf                 = ws.cell(row=r, column=4).value
        stored_horizon     = ws.cell(row=r, column=5).value
        last_predicted_val = ws.cell(row=r, column=6).value

        # Read forecasted sequence from ForecastDetails sheet
        ws_detail = wb["Forecast Details"]
        detail_headers = [h.value for h in ws_detail[1]]
        forecasted_prices_col = detail_headers.index("ForecastedPrices") + 1
        forecasted_seq_val = None
        for dr in range(2, ws_detail.max_row + 1):
            if ws_detail.cell(row=dr, column=1).value == row_id:
                forecasted_seq_val = ws_detail.cell(row=dr, column=forecasted_prices_col).value
                break

        try:
            forecast_time = pd.to_datetime(forecast_time_val)
            horizon = int(stored_horizon)
            last_predicted = float(last_predicted_val)
            forecasted = np.array([float(v) for v in forecasted_seq_val.split(",")]) if forecasted_seq_val else None
        except Exception as e:
            logging.info(f"  Skipping row {r} – parse error: {e}")
            return {}

        tf = str(tf)

        # Normalize frequency for pandas (m -> min for minute, h -> h for hour)
        pandas_freq = re.sub(r'(?i)(\d+)m$', r'\1min', tf)
        # Round ForecastTime down to the interval boundary
        if forecast_time.tz is not None:
            forecast_time = forecast_time.tz_localize(None)
        rounded_ft = forecast_time.floor(pandas_freq)

        # Compute each forecast step timestamp
        step_duration = pd.to_timedelta(tf)
        step_times = [rounded_ft + i * step_duration for i in range(1, horizon + 1)]

        logging.info(f"  ForecastTime:    {forecast_time}")
        logging.info(f"  Rounded to:      {rounded_ft}")
        logging.info(f"  Target windows:  {step_times[0]} to {step_times[-1]}")

        # Extract close prices from downloaded data
        close = data["Close"]
        if isinstance(close, pd.DataFrame):
            close = close.iloc[:, 0]
        close = close.dropna()
        if close.index.tz is not None:
            close.index = close.index.tz_localize(None)

        logging.info(f"  Data range:      {close.index[0]} to {close.index[-1]}")

        # Look up each step timestamp in the data
        actual_vals = []
        for ts in step_times:
            idxs = close.index.get_indexer([ts], method='nearest')
            if idxs[0] != -1:
                actual_vals.append(float(close.iloc[idxs[0]]))
            else:
                logging.info(f"  -> Warning: No data found near {ts}.")
                return {}

        actual_array = np.array(actual_vals)

        last_actual   = float(actual_array[-1])
        abs_error     = abs(last_predicted - last_actual)
        pct_error     = (abs_error / last_actual * 100) if last_actual != 0 else 0.0
        mape          = float(np.mean(np.abs((actual_array - forecasted) / actual_array)) * 100) if np.all(actual_array != 0) else 0.0

        ws.cell(row=r, column=7,  value=last_actual)
        ws.cell(row=r, column=8,  value=abs_error)
        ws.cell(row=r, column=9,  value=pct_error)
        ws.cell(row=r, column=10, value=mape)

        # Update Forecast Details sheet with actuals and metrics
        ws_detail = wb["Forecast Details"]
        detail_headers = [h.value for h in ws_detail[1]]
        actual_prices_str = ",".join(f"{v:.6f}" for v in actual_array)
        abs_error_str = ",".join(f"{abs(f - a):.6f}" for f, a in zip(forecasted, actual_array))
        pct_error_str = ",".join(f"{abs(f - a) / a * 100:.6f}" if a != 0 else "0" for f, a in zip(forecasted, actual_array))

        # Find the matching row in ForecastDetails by ID
        for detail_r in range(2, ws_detail.max_row + 1):
            if ws_detail.cell(row=detail_r, column=1).value == row_id:
                ws_detail.cell(row=detail_r, column=detail_headers.index("ActualPrices") + 1, value=actual_prices_str)
                ws_detail.cell(row=detail_r, column=detail_headers.index("AbsError") + 1, value=abs_error_str)
                ws_detail.cell(row=detail_r, column=detail_headers.index("PctError") + 1, value=pct_error_str)
                ws_detail.cell(row=detail_r, column=detail_headers.index("MAPE") + 1, value=mape)
                logging.info(f"  -> Updated ForecastDetails row {detail_r}.")
                break

        # Chart generation
        hist = close.loc[:rounded_ft]
        hist_idx = hist.index[-200:]
        hist_prc = hist.values[-200:]

        plt.figure(figsize=(7, 3.5))
        plt.plot(hist_idx, hist_prc, label="Historical")
        plt.plot(step_times, forecasted, label="Forecast")
        plt.plot(step_times, actual_array, label="Actual_Price")
        plt.legend(fontsize=8)
        plt.grid(True)
        plt.title(f"{coin} {tf} Evaluation – ID: {row_id}", fontsize=9)
        plt.tight_layout()

        chart_time_str = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        chart_filename = f"chart_{row_id}_{chart_time_str}.png"
        chart_path = os.path.join(self.charts_dir, chart_filename)
        plt.savefig(chart_path, bbox_inches='tight', dpi=100)
        plt.close()

        # try:
        #     img = OpenpyxlImage(chart_path)
        #     img.width  = int(img.width  * 0.5)
        #     img.height = int(img.height * 0.5)
        #     ws.row_dimensions[r].height = 130
        #     ws.add_image(img, f"{IMAGE_COL}{r}")
        #     logging.info(f"  -> Chart saved and embedded.")
        # except Exception as e:
        #     logging.info(f"  -> Warning: Could not embed chart: {e}")

        wb.save(self.log_filepath)
        logging.info(f"  -> Metrics: last_actual={last_actual:.2f}, abs_error={abs_error:.2f}, "
              f"pct_error={pct_error:.2f}%, mape={mape:.2f}%")
        return {
            "evaluated": True,
            "row_id": row_id,
            "last_predicted": last_predicted,
            "last_actual": last_actual,
            "mape": mape,
            "chart_path": chart_path,
        }
